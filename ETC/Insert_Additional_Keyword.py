# -*- coding: utf-8 -*-

import openpyxl

from tkinter import *
from tkinter import filedialog


root = Tk()
root.withdraw()
root.filename = filedialog.askopenfilename(initialdir="D:/Dropbox/A-TEB/작업폴더/", title="Open Data files", filetypes=(("data files", "*.csv;*.xls;*.xlsx"), ("all files", "*.*")))

wb = openpyxl.load_workbook(root.filename)
sheet = wb.worksheets[0]
Max_row = sheet.max_row
chaged = 0
target_text = "</FONT></CENTER>"

final_detail_page_column = 49
additional_keywords_column = 108

for row_ptr in range(2, Max_row + 1):
    original_html = sheet.cell(row=row_ptr, column=final_detail_page_column).value
    additional_keywords = sheet.cell(row=row_ptr, column=additional_keywords_column).value
    if additional_keywords == "" or additional_keywords == None:
        continue
    else:
        sheet.cell(row=row_ptr, column=49).value = original_html.replace(target_text, " " + additional_keywords + " " + target_text)
        chaged += 1

print("Total : {0} \nChanged : {1}".format(Max_row - 1, chaged))

wb.save(root.filename)