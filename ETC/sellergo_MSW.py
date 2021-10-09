# -*- coding: utf-8 -*-

import openpyxl

from tkinter import *
from tkinter import filedialog


root = Tk()
root.withdraw()
root.filename = filedialog.askopenfilename(initialdir="D:\\Dropbox\\A-TEB\\작업폴더\\99.옵션품절\\", title="Open Data files", filetypes=(("data files", "*.csv;*.xls;*.xlsx"), ("all files", "*.*")))

wb = openpyxl.load_workbook(root.filename)
sheet = wb.worksheets[0]
# MaxOptionCount = int(input("Max Option Number : "))
MaxOptionCount = 10
Offset = 4
Blank = 2

# SOcode = ""
# Checked = 0

# for RowPointer in range(2, sheet.max_row+1):
#     Options = sheet.cell(row=RowPointer, column=3).value
#     OptionsCount = Options.count('**')
#     if MaxOptionCount < OptionsCount:
#         MaxOptionCount = OptionsCount
#     for i in range(0, MaxOptionCount+1):
#         sheet.cell(row=RowPointer, column=5+i).value = ""




# for RowPointer in range(2, sheet.max_row+1):
    #
    # Options = sheet.cell(row=RowPointer, column=3).value
    # OptionsCount = Options.count('**')
    # OptionList = Options.split('**')
    #
    # for j in range(OptionsCount):
    #
    #     Option = OptionList[j]
    #     ClearedOption = Option.split('*')
    #     ActualOption = ClearedOption[1]
    #
    #     if ActualOption.find(':') > 0:
    #         opt1 = ActualOption.split(':')[0]
    #         opt2 = ActualOption.split(':')[1]
    #     else:
    #         opt1 = ActualOption
    #         opt2 = ''
    #
    #     SOcode = sheet.cell(row=RowPointer, column=2).value + " " + opt1 + " " + opt2
    #     SOcode = SOcode.replace("(", " ")
    #     SOcode = SOcode.replace(")", " ")
    #     SOcode = SOcode.replace("~", " ")
    #     SOcode = SOcode.replace("/", " ")
    #     sheet.cell(row=RowPointer, column= 5+j).value = SOcode


# 원기록 복사

for i in range(2, sheet.max_row+1):
    for j in range(MaxOptionCount+1):
        sheet.cell(row=i, column=MaxOptionCount+j+Offset+Blank).value = sheet.cell(row=i, column=j+Offset).value


# 중복옵션 지우기

for i in range(2, sheet.max_row+1):
    # sheet.cell(row=i, column=4).value = " "

    for j in range(MaxOptionCount+Offset+Blank, MaxOptionCount+Offset+Blank+MaxOptionCount):
        for k in range(MaxOptionCount+Offset+Blank, MaxOptionCount+Offset+Blank+MaxOptionCount):
            if sheet.cell(row=i, column=j).value == sheet.cell(row=i+1, column=k).value:
                sheet.cell(row=i, column=j).value = ""
                sheet.cell(row=i+1, column=k).value = ""
                break
            elif sheet.cell(row=i, column=j).value == None or sheet.cell(row=i, column=j).value == "" :
                break
            # else:
            #     sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("(", " ")
            #     sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace(")", " ")
            #     sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("/", " ")
            #     sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("~", " ")

for i in range(2, sheet.max_row+1):
    for j in range(MaxOptionCount+Offset+Blank, MaxOptionCount+Offset+Blank+MaxOptionCount):
        if sheet.cell(row=i, column=j).value == None or sheet.cell(row=i, column=j).value == "" :
            sheet.cell(row=i, column=j).value == ""
        else:
            sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("(", " ")
            sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace(")", " ")
            sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("/", " ")
            sheet.cell(row=i, column=j).value = str(sheet.cell(row=i, column=j).value).replace("~", " ")



# for i in range(2, sheet.max_row+1):
#     for j in range(5, MaxOptionCount+5):
#         if sheet.cell(row=i, column=j) != "":
#             Checked += 1
#
# print("Checked : ", Checked)


wb.save(root.filename)






