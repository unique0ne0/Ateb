# -*- coding: utf-8 -*-

import openpyxl

from tkinter import *
from tkinter import filedialog


# Open Excel File
root = Tk()
root.withdraw()
root.filename = filedialog.askopenfilename(initialdir="D:\Dropbox\A-TEB\작업폴더\99.옵션품절/", title="Open Data files", filetypes=(("data files", "*.csv;*.xls;*.xlsx"), ("all files", "*.*")))

wb = openpyxl.load_workbook(root.filename)

# Set Variables

try:
    del wb['SO']
    del wb['OSO']
except:
    pass

wb.create_sheet('SO', 2)
wb.create_sheet('OSO', 3)

Old = wb.worksheets[0]
New = wb.worksheets[1]

SO = wb.worksheets[2]
OSO = wb.worksheets[3]

OldMaxRow = Old.max_row
NewMaxRow = New.max_row

SO.cell(row=1, column=1).value = '상품코드'
SO.cell(row=1, column=2).value = '변경상품명'
SO.column_dimensions['A'].width = 20

OSO.cell(row=1, column=1).value = '상품코드'
OSO.cell(row=1, column=2).value = '변경상품명'
OSO.cell(row=1, column=3).value = '품절옵션'
OSO.column_dimensions['A'].width = 20
OSO.column_dimensions['B'].width = 40

SORowPointer = 2
OSORowPointer = 2


# Routine 1 - Search Duplicated Products
for OldRowPointer in range(2, OldMaxRow+1):

    OriginalPN = Old.cell(row= OldRowPointer, column=2).value
    IsDuplicated = 0
    ColumnPointer = 3

    for NewRowPointer in range(2, NewMaxRow+1):

        NewPN = New.cell(row= NewRowPointer, column=2).value

        # 이전 수집 내용과 신규 수집 내용 비교하여 중복되면 중복 표시 및 옵션 검토
        if OriginalPN == NewPN:
            IsDuplicated = 1

            OldOptions = Old.cell(row= OldRowPointer, column=30).value
            OldOptionCount = OldOptions.count('**')
            OldOptionList = OldOptions.splitlines()

            NewOptions = New.cell(row=NewRowPointer, column=30).value
            NewOptionCount = NewOptions.count('**')
            NewOptionList = NewOptions.splitlines()

            ColumnPointer = 3
            OSOChanged = 0

            for j in range(OldOptionCount):

                InOption = 0

                OldOption = OldOptionList[j]
                OldOptionContents = OldOption.split('*')
                Cleared_OldOption = OldOptionContents[0] + " " + OldOptionContents[1]

                # 신규 수집 옵션 비교해서 존재하면 품절 아닌 것으로 판정
                for k in range(NewOptionCount):
                    NewOption = NewOptionList[k]
                    NewOptionContents = NewOption.split('*')
                    Cleared_NewOption = NewOptionContents[0] + " " + NewOptionContents[1]

                    if Cleared_OldOption == Cleared_NewOption:
                        InOption = 1
                        break

                # 옵션내용이 누락되면 옵션품절로 판정 후 기록
                if InOption == 0:
                    OSO.cell(row=OSORowPointer, column=1).value = OriginalPN
                    OSO.cell(row=OSORowPointer, column=2).value = Old.cell(row=OldRowPointer, column=4).value
                    TargetText = OriginalPN + " " + Cleared_OldOption
                    OSO.cell(row=OSORowPointer, column=ColumnPointer).value = TargetText.replace("/"," ")
                    OSO.column_dimensions[OSO.cell(row=OSORowPointer, column=ColumnPointer).column_letter].width = 40
                    ColumnPointer += 1
                    OSOChanged = 1

            if OSOChanged == 1:
                OSORowPointer += 1
                OSOChanged = 0

    # 이전 수집 내용과 신규 수집 내용 비교하여 중복되지 않으면 품절로 체크
    if IsDuplicated == 0:
        SO.cell(row=SORowPointer, column=1).value = Old.cell(row=OldRowPointer, column=2).value
        SO.cell(row=SORowPointer, column=2).value = Old.cell(row=OldRowPointer, column=4).value
        SORowPointer += 1



print("SO Changed : {0} \nOSO Changed : {1}".format(SORowPointer-2, OSORowPointer-2))

wb.save(root.filename)




