# -*- coding: utf-8 -*-

import sys
import hashlib
import hmac
import ssl
import time
import urllib.request
import json

import os
from datetime import datetime, date, timedelta

# tkinter import 주의 : timezone 적용이 안 됨.
from tkinter import *
from tkinter import filedialog

import openpyxl

import logging

# --------------------------------------------------
# Split file Name : CP_API_Variables_configuration.py
# Start---------------------------------------------

# Variables Setting
# """
# 기본 변수 설정
#
# os.environ['TZ'] = TimeZone 설정
# vendorID = 쿠팡 판매자 ID
# accesskey = 쿠팡 API accesskey
# secretkey = 쿠팡 API secretkey
# contractID = 쿠팡 계약 번호
# default_discount_rate = 디폴트 할인율
# default_coupon_period = 디폴트 쿠폰 적용 기간
# default_Max_discount_price = 디폴트 최대 할인금액
# default_Min_price = 디폴트 할인시작 최소 금액
# userID = 쿠팡윙 사용자 ID
# n_Max_coupon_P_daily = 1일 최대 쿠폰 수
# n_Max_item =  1회 최대 처리 엑셀 행(옵션) 수
# """

os.environ['TZ'] = 'GMT+0'
vendorID = "A00197644"
accesskey = "9a070fbb-9c53-45e0-9096-b6367f5cef00"
secretkey = "cbdda91a98dd734a05b1c25a9f5cbc093f0e9b44"
contractID = 61553
default_discount_rate = 2
default_coupon_period = 180
default_Max_discount_price = 3000
default_Min_price = 20000
userID = "marketjnj"
n_Max_coupon_P_daily = 1
# n_Max_item = 10000
modulename = 'tkinter'
maximumPerDaily = 1

# End-----------------------------------------------
# Split file Name : Variables Setting.py
# --------------------------------------------------



# --------------------------------------------------
# Split file Name : json_handler.py
# Start---------------------------------------------

class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (bytes, bytearray)):
            return obj.decode("utf8") # <- or any other encoding of your choice
        # Let the base class default method raise the TypeError
        return json.JSONEncoder.default(self, obj)

# End-----------------------------------------------
# Split file Name : json_handler.py
# --------------------------------------------------

# --------------------------------------------------
# Split file Name : logger_handler.py
# Start---------------------------------------------

def set_Logger(name=None):
    # 1 logger instance를 만든다.
    logger = logging.getLogger(name)

    # 2 logger의 level을 가장 낮은 수준인 DEBUG로 설정해둔다.
    logger.setLevel(logging.INFO)

    # 3 formatter 지정
    formatter = logging.Formatter("\n----- %(asctime)s / %(funcName)s / %(levelname)s : (%(lineno)d) %(message)s")

    # 4 handler instance 생성
    console = logging.StreamHandler()
    file_handler = logging.FileHandler(filename="test.log")

    # 5 handler 별로 다른 level 설정
    console.setLevel(logging.INFO)
    # console.setLevel(logging.WARNING)
    # file_handler.setLevel(logging.DEBUG)

    # 6 handler 출력 format 지정
    console.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # 7 logger에 handler 추가
    logger.addHandler(console)
    # logger.addHandler(file_handler)

    return logger


# End-----------------------------------------------
# Split file Name : logger_handler.py
# --------------------------------------------------


# --------------------------------------------------
# Split file Name : Excel_handler.py
# Start---------------------------------------------

def open_ExcelFile():
    """
    엑셀파일 열기
    :return: 파일명 리턴
    """
    logger.info("Function In")

    initial_directory = "./"

    root = Tk()
    root.withdraw()
    root.filename = filedialog.askopenfilename(initialdir=initial_directory, title="Open Data files",
                                               filetypes=(("data files", "*.csv;*.xls;*.xlsx"), ("all files", "*.*")))

    return root.filename

# End-----------------------------------------------
# Split file Name : Excel_handler.py
# --------------------------------------------------

def load_CPExcel(API_mode):
    # """
    # 엑셀파일 로딩후 API 모드에 따라 대상 ID 리스트 및 갯수 확인
    # :param API_mode: API 모드
    # :return: ID 리스트 갯수, ID 리스트의 리스트
    # """

    logger.info("Function In")

    start_row = 4

    if API_mode == "IDC_Create" or API_mode == "DLC_Create":
        target_ID_column = 3
    elif API_mode == "ITEM_Delete":
        target_ID_column = 1

    xlsx_filename = open_ExcelFile()
    wb = openpyxl.load_workbook(xlsx_filename, data_only=True)
    ws = wb.active

    total_rows = ws.max_row - start_row + 1
    if API_mode == "IDC_Create":
        n_Max_item = 10000
    elif API_mode == "DLC_Create":
        n_Max_item = 1000
    elif API_mode == "ITEM_Delete":
        n_Max_item = 1000000

    n_target_ID_list = int(total_rows / n_Max_item) + 1
    target_ID_list = list()

    for i in range(n_target_ID_list):

        rows_ptr = i * n_Max_item + start_row

        if i == n_target_ID_list - 1:
            rng_rows = total_rows - (n_target_ID_list - 1) * n_Max_item
        else:
            rng_rows = n_Max_item

        temp_list = list()

        for row in range(rows_ptr, rows_ptr + rng_rows):
            temp_list.append(ws.cell(row=row, column=target_ID_column).value)

        target_ID_list.append(temp_list)

    ID_list = dict(n_target_ID_list=n_target_ID_list, target_ID_list=target_ID_list)
    # """
    # n_target_ID_list : ID 리스트 갯수
    # target_ID_list : ID 리스트의 리스트
    # """

    return ID_list


def get_API_mode():
    # """
    # 작업모드 선택
    # :return: API 모드
    # """

    logger.info("Function In")

    while True:

        print("======== 쿠팡 API ========")
        print("1. 즉시할인 쿠폰 생성 및 적용")
        print("2. 다운로드 쿠폰 생성 및 적용")
        print("3. 판매중지 상품 삭제")
        print("0. 종료")

        mode = int(input("Coupon type? : "))

        if mode == 1:
            return "IDC_Create"
        elif mode == 2:
            return "DLC_Create"
        elif mode == 3:
            return "ITEM_Delete"
        elif mode == 0:
            exit()


# --------------------------------------------------
# Split file Name : CP_API_handler.py
# Start---------------------------------------------


def API_Setting(*args):
    # """
    # API 모드에 따라 method, path, message, url, authorization 처리
    # :param args: API 모드(필수), transferedID or sellerProductID(옵션)
    # :return: preAPI: dict of {method, path, message, url, authorization}
    # """

    logger.info("Function In")

    API_mode = args[0]

    if API_mode == "IDC_Create":
        method = "POST"
        path = f"/v2/providers/fms/apis/api/v2/vendors/{vendorID}/coupon"
    elif API_mode == "Check_IDC_Coupon":
        method = "GET"
        transferedID = args[1]
        path = f"/v2/providers/fms/apis/api/v1/vendors/{vendorID}/requested/{transferedID}"
    elif API_mode == "IDC_ITEM_Create":
        method = "POST"
        transferedID = args[1]
        path = f"/v2/providers/fms/apis/api/v1/vendors/{vendorID}/coupons/{transferedID}/items"
    elif API_mode == "ITEM_Delete":
        method = "DELETE"
        sellerProductID = args[1]
        path = f"/v2/providers/seller_api/apis/api/v1/marketplace/seller-products/{sellerProductID}"
    elif API_mode == "DLC_Create":
        method = "POST"
        path = "/v2/providers/marketplace_openapi/apis/api/v1/coupons"
    elif API_mode == "DLC_ITEM_Create":
        method = "PUT"
        path = "/v2/providers/marketplace_openapi/apis/api/v1/coupon-items"

    if modulename in sys.modules:                   # tkinter 임포트 여부에 따라 타임존 설정
        GMT0 = datetime.now() - timedelta(hours=9)
        date_time = time.strftime('%y%m%d') + 'T' + GMT0.strftime('%H%M%S') + 'Z'
    else:
        date_time = time.strftime('%y%m%d') + 'T' + time.strftime('%H%M%S')+'Z'

    message = date_time + method + path
    url = "https://api-gateway.coupang.com" + path

    signature = hmac.new(secretkey.encode('utf-8'), message.encode('utf-8'), hashlib.sha256).hexdigest()
    authorization = "CEA algorithm=HmacSHA256, access-key=" + accesskey + ", signed-date=" + date_time + ", signature=" + signature

    return dict(method=method, path=path, message=message, url=url, authorization=authorization)


def send_Request(method, url, authorization, str_json):
    # """
    # 쿠팡 API 호출
    # :param method:
    # :param url:
    # :param authorization:
    # :param str_json:
    # :return:
    # """

    logger.info("Function In")

    logger.info(method)
    logger.info(url)
    logger.info(authorization)
    logger.info(json.dumps(str_json, indent=2, ensure_ascii=False, cls=MyEncoder))

    print(f"\n************* SEND {method} REQUEST *************")

    req = urllib.request.Request(url)

    req.add_header("Content-type", "application/json;charset=UTF-8")
    req.add_header("Authorization", authorization)

    req.get_method = lambda: method

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    try:
        if method == "POST" or method == "PUT":
            resp = urllib.request.urlopen(req, str_json, context=ctx)

        elif method == "GET" or method == "DELETE":
            resp = urllib.request.urlopen(req, context=ctx)

    except urllib.request.HTTPError as e:
        print(e.code)
        print(e.reason)

        # print(json.dumps(json.loads(e.fp.read()), ensure_ascii=False))
        print(json.dumps(json.loads(e.fp.read()), indent=2, sort_keys=False, ensure_ascii=False))

    except urllib.request.URLError as e:
        print(e.errno)
        print(e.reason)

        # print(json.dumps(json.loads(e.fp.read()), ensure_ascii=False))
        print(json.dumps(json.loads(e.fp.read()), indent=2, sort_keys=False, ensure_ascii=False))


    else:
        # 200
        body = resp.read().decode(resp.headers.get_content_charset())
        # print(json.dumps(json.loads(body), ensure_ascii=False))
        print(json.dumps(json.loads(body), indent=2, sort_keys=False, ensure_ascii=False))
        print("\n\n")

        return body

# END-----------------------------------------------
# Split file Name : CP_API_handler.py
# --------------------------------------------------


def get_IDC_RequestBodyList(n_OptionID, discount_rate, coupon_period, Max_discount_price):
    # """
    # 할인쿠폰 생성을 위한 Request Body Parameter List 생성
    # :param n_OptionID: 옵션ID 리스트 갯수
    # :param discount_rate: 할인율
    # :param coupon_period: 적용기간
    # :param Max_discount_price: 최대할인금액
    # :return: IDC_request_body_list: Request Body List
    # """

    logger.info("Function In")

    IDC_request_body_list = list()

    for i in range(n_OptionID):
        name = date.today().strftime("%m") + "월 고객감사 즉시 할인쿠폰 " + str(i+1)
        startAt = (date.today() + timedelta(1)).strftime("%Y-%m-%d ") + "00:00:00"
        endAt = (date.today() + timedelta(coupon_period)).strftime("%Y-%m-%d ") + "00:00:00"
        request_body = dict(contractId=contractID, name=name, maxDiscountPrice=Max_discount_price, discount=discount_rate,
                           startAt=startAt, endAt=endAt, type="RATE", wowExclusive=False)
        IDC_request_body_list.append(request_body)

    logging.info(json.dumps(IDC_request_body_list, indent=2, ensure_ascii=False))

    return IDC_request_body_list


def create_IDC_RequestIDList(preAPI, IDC_request_body_list):
    # """
    # 즉시할인쿠폰 생성을 위한 RequestedID 리스트 생성
    # :param preAPI: dict of {method, path, message, url, authorization}
    # :param IDC_request_body_list: Request Body 리스트
    # :return: requestID_list: RequestID 리스트
    # """

    logger.info("Function In")

    # loop - send request
    requestID_list = list()

    for i in range(len(IDC_request_body_list)):

        logging.info(f"POST Request ID request : {i+1}")

        body_parameter = json.dumps(IDC_request_body_list[i], cls=MyEncoder).encode("utf-8")

        requestID_list.append(json.loads(send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], body_parameter))['data']['content']['requestedId'])

    return requestID_list


def create_IDC_CouponIDList(IDC_requestID_list):
    # """
    # 즉시할인 쿠폰ID 리스트 생성
    # :param IDC_requestID_list: 쿠폰ID 생성을 위한 requestID 리스트
    # :return: IDC_couponID_list: 쿠폰ID 리스트
    # """

    logger.info("Function In")

    IDC_couponID_list = list()

    # loop - send request
    for i in range(len(IDC_requestID_list)):

        logging.info(f"Get IDC CouponID Request : {i}")

        requestedID = IDC_requestID_list[i]


        preAPI = API_Setting("Check_IDC_Coupon", requestedID) # return : dict of {method, path, message, url, authorization}

        IDC_couponID_list.append(json.loads(send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], requestedID))['data']['content']['couponId'])

    return IDC_couponID_list


def create_IDC(API_mode):
    # """
    # 즉시할인쿠폰 생성
    # :param API_mode: API 모드
    # :return: 즉시할인쿠폰 리스트, 대상 옵션ID 리스트
    # """

    logger.info("Function In")

    ID_list = list()

    print("\n===== 즉시할인 쿠폰 생성 및 적용 =====\n")

    # 적용 엑셀 파일 선택
    print("대상 엑셀 파일 선택\n")
    ID_list = load_CPExcel(API_mode)                    # 엑셀파일 로딩 및 대상 ID 리스트 확인
    """
    ID_list : dict of {n_target_ID_list(리스트 갯수), target_ID_list(대상ID 리스트)}
    """

    # 적용 할인율 선택
    input_rate = int(input("Input Discount Rate (0=cancel/1=default/2~=Discount Rate) : "))

    if input_rate == 0:
        # 0: 작업종료
        IDC_couponID_list = list()
        ID_list['n_target_ID_list'] = 0
        ID_list['target_ID_list'] = list()
        pass
    elif input_rate == 1:
        # 1 : 디폴트 설정 적용
        print("디폴트는 2%, 날짜는 내일부터 180일간, 최대 할인값은 3000원")
        discount_rate = default_discount_rate
        coupon_period = default_coupon_period
        Max_discount_price = default_Max_discount_price
    else:
        discount_rate = input_rate
        coupon_period = int(input("Coupon Period : "))
        Max_discount_price = int(input("Max Discount Price : "))

    preAPI = API_Setting("IDC_Create") # return : dict of {method, path, message, url, authorization}

    # request list 만들기
    IDC_request_body_list = get_IDC_RequestBodyList(ID_list['n_target_ID_list'], discount_rate, coupon_period, Max_discount_price)

    # request list에서 requestID list 만들기
    IDC_requestID_list = create_IDC_RequestIDList(preAPI, IDC_request_body_list)

    # requestID list로 즉시할인쿠폰 생성
    IDC_couponID_list = create_IDC_CouponIDList(IDC_requestID_list)

    return {'IDC_couponID_list': IDC_couponID_list, 'optionID_list': ID_list['target_ID_list']}


def apply_IDC(IDC):
    # """
    # 즉시할인 쿠폰을 각 옵션에 적용
    # :param IDC: dict of IDC_couponID_list, optionID_list
    # :param IDC_couponID_list: 즉시할인 쿠폰ID 리스트
    # :param optionID_list: 쿠폰 적용 대상 옵션ID 리스트
    # :return: apply_requestedID_list: 적용된 ID 리스트
    # """

    logger.info("Function In")

    if len(IDC['IDC_couponID_list']) == 0:
        return []

    apply_requestedID_list = list()

    for i in range(len(IDC['IDC_couponID_list'])):

        preAPI = API_Setting("IDC_ITEM_Create", IDC['IDC_couponID_list'][i])

        logger.info(f"POST Coupon Apply Request : {i}")

        IDC_apply_coupon_item_list = dict(vendorItems=IDC['optionID_list'][i])

        body_parameter = json.dumps(IDC_apply_coupon_item_list, cls=MyEncoder).encode("utf-8")

        logger.info(json.dumps(IDC_apply_coupon_item_list, indent=2, ensure_ascii=False))

        apply_requestedID_list.append(json.loads(send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], body_parameter))['data']['content']['requestedId'])

    return apply_requestedID_list


def get_DLC_RequestBodyList(n_OptionID, policies):
    # """
    # 다운로드 쿠폰 생성을 위한 Request Body List 생성
    # :param n_OptionID: 옵션ID 리스트 갯수
    # :param coupon_policy: 쿠폰 정책 dict
    # :return: DLC_request_body_list: Request Body List
    # """

    logger.info("Function In")

    DLC_request_body_list = list()

    for i in range(n_OptionID):

        main_coupon_title = date.today().strftime("%m") + "월 고객감사 시즌ON 쿠폰"
        startDate = (date.today() + timedelta(1)).strftime("%Y-%m-%d ") + "00:00:00"
        endDate = (date.today() + timedelta(default_coupon_period)).strftime("%Y-%m-%d ") + "00:00:00"
        request_body = dict(title=main_coupon_title, contractId=contractID, couponType="DOWNLOAD", startDate=startDate, \
                            endDate=endDate, userId=userID, policies=policies)
        DLC_request_body_list.append(request_body)

    logging.info(json.dumps(DLC_request_body_list, indent=2, ensure_ascii=False))

    return DLC_request_body_list


def create_DLC_CouponIDList(DLC_request_body_list):
    # """
    # 다운로드 쿠폰ID 리스트 생성
    # :param DLC_request_body_list: 쿠폰ID 생성을 위한 request body 리스트
    # :return: DLC_couponID_list: 쿠폰ID 리스트
    # """

    logger.info("Function In")

    DLC_couponID_list = list()

    # loop - send request
    for i in range(len(DLC_request_body_list)):

        logging.info(f"Get DLC CouponID Request : {i}")

        preAPI = API_Setting("DLC_Create") # return : dict of {method, path, message, url, authorization}

        logger.info(json.dumps(DLC_request_body_list, indent=2, ensure_ascii=False))

        body_parameter = json.dumps(DLC_request_body_list[i], cls=MyEncoder).encode("utf-8")

        DLC_couponID_list.append(json.loads(send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], body_parameter))['couponId'])

    return DLC_couponID_list


def create_DLC(API_mode):
    # """
    # 다운로드쿠폰 생성
    # :param API_mode: API 모드
    # :return: 다운로드쿠폰 리스트, 대상 옵션ID 리스트
    # """
    logger.info("Function In")

    ID_list = list()
    policies = []

    print("\n===== 다운로드 쿠폰 생성 및 적용 =====\n")

    # 적용 엑셀 파일 선택
    print("대상 엑셀 파일 선택\n")
    ID_list = load_CPExcel(API_mode)  # 엑셀파일 로딩 및 대상 ID 리스트 확인
    """
    ID_list : dict of {n_target_ID_list(리스트 갯수), target_ID_list(대상ID 리스트)}
    """

    # 적용 정책 설정 _ 디폴트 적용
    print("\n----- 디폴트 설정 적용 -----\n\n")
    print("-- 쿠폰명 : 8888 시즌 ON 쿠폰 할인 \n")
    print("-- 쿠폰유효기간 : 발행일로부터 30일 \n")
    print("-- 할인방식 : 정액 15000, 500/ 30000, 1000/ 45000, 1500 \n")
    print("-- 최대발급개수 : 1 \n")

    main_coupon_title = "8888 고객감사 시즌 ON 할인 쿠폰"
    default_minimumPrice1 = 15000
    default_discount1 = 500
    default_minimumPrice2 = 30000
    default_discount2 = 1000
    default_minimumPrice3 = 45000
    default_discount3 = 1500

    coupon_policy1 = dict(title='구매감사 할인쿠폰', typeOfDiscount="PRICE", description='구매감사 할인쿠폰',
                          minimumPrice=default_minimumPrice1, discount=default_discount1,
                          maximumDiscountPrice=0, maximumPerDaily=maximumPerDaily)
    coupon_policy2 = dict(title='구매감사 할인쿠폰', typeOfDiscount="PRICE", description='구매감사 할인쿠폰',
                          minimumPrice=default_minimumPrice2, discount=default_discount2,
                          maximumDiscountPrice=0, maximumPerDaily=maximumPerDaily)
    coupon_policy3 = dict(title='구매감사 할인쿠폰', typeOfDiscount="PRICE", description='구매감사 할인쿠폰',
                          minimumPrice=default_minimumPrice3, discount=default_discount3,
                          maximumDiscountPrice=0, maximumPerDaily=maximumPerDaily)

    policies.append(coupon_policy1)
    policies.append(coupon_policy2)
    policies.append(coupon_policy3)

    # request body list 작성
    DLC_request_body_list = get_DLC_RequestBodyList(ID_list['n_target_ID_list'], policies)

    # request body list로 다운로드 쿠폰ID 리스트 생성
    DLC_couponID_list = create_DLC_CouponIDList(DLC_request_body_list)

    return {'DLC_couponID_list': DLC_couponID_list, 'optionID_list': ID_list['target_ID_list']}


def apply_DLC(DLC):
    # """
    # 다운로드 쿠폰을 각 옵션에 적용
    # :param DLC: dict of DLC_couponID_list, optionID_list
    # :param DLC_couponID_list: 다운로드 쿠폰ID 리스트
    # :param optionID_list: 쿠폰 적용 대상 옵션ID 리스트
    # :return: applied_ID_list: 적용된 ID 리스트
    # """

    logger.info("Function In")

    if len(DLC['DLC_couponID_list']) == 0:
        return []

    applied_couponID_list = []

    for i in range(len(DLC['DLC_couponID_list'])):

        preAPI = API_Setting("DLC_ITEM_Create")

        logger.info(f"PUT DLC Coupon Apply Request : {i}")

        couponItem_dict = {}
        couponItem_list = []
        couponItems_dict = {}
        couponItems = []


        couponItem_dict = dict(couponId=DLC['DLC_couponID_list'][i], userId=userID, vendorItemIds=DLC['optionID_list'][i])
        couponItem_list.append(couponItem_dict)

        couponItems = dict(couponItems=couponItem_list)

        body_parameter = json.dumps(couponItems, cls=MyEncoder).encode("utf-8")

        logger.info(json.dumps(couponItems, indent=2, ensure_ascii=False))

        applied_couponID_list.append(json.loads(send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], body_parameter)))

    return applied_couponID_list


def delete_Item(API_mode):
    # """
    # 판매중지 상품 삭제
    # :param API_mode: API 모드
    # :return:
    # """

    logger.info("Function In")

    # 적용 엑셀 파일 선택
    print("대상 엑셀 파일 선택\n")
    ID_list = load_CPExcel(API_mode)['target_ID_list'][0]                    # 엑셀파일 로딩 및 대상 ID 리스트 확인

    for i in range(len(ID_list)):

        sellerProductID =  ID_list[i]

        preAPI = API_Setting("ITEM_Delete", sellerProductID)

        send_Request(preAPI['method'], preAPI['url'], preAPI['authorization'], ID_list[i])


def main():

    while True:

        API_mode = get_API_mode()             # 모드 확인

        if API_mode == "IDC_Create":         # 즉시할인쿠폰 모드

            IDC = create_IDC(API_mode)       # 즉시할인쿠폰 생성

            appliedIDC = apply_IDC(IDC)      # 즉시할인쿠폰 적용

            print(appliedIDC)

        elif API_mode == "DLC_Create":       # 다운로드 쿠폰 모드

            DLC = create_DLC(API_mode)       # 다운로드 쿠폰 생성

            appliedDLC = apply_DLC(DLC)      # 다운로드 쿠폰 적용

            print(appliedDLC)

        elif API_mode == "ITEM_Delete":      # 판매중지상품 삭제 모드

            delete_Item(API_mode)

        else: sys.exit()



if __name__ == '__main__':

    # 로거 설정 및 생성
    logger = set_Logger()

    main()

