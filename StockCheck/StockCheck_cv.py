# -*- coding: utf-8 -*-

import time
from tkinter import *
from tkinter import filedialog

import logging
import openpyxl
import pyexcel
import shutil
from bs4 import BeautifulSoup
from colored import fg, attr
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from win32com.client import Dispatch
import re

# 변수 설정

email_id = "marketjnj@gmail.com"
# pw = "1q2wazsx!!"
pw = "12345678p"
target_url = "http://app.sellergo.net"
default_Max_wait_seconds = 2
NOT_CHECKED = 0
CHECKED = 1
SLG_OSO = 1
SELL = 2
DEFAULT_STOCK = 9
remove_char = str.maketrans(":~()_-/[].／", "           ")

HEADER_ROW = 1 - 1
PRODUCT_CODE_COLUMN = 2 - 1
PRODUCT_NAME_COLUMN = 4 - 1
OPTION_INFO_COLUMN = 19 - 1
SOLD_AMNT_COLUMN = 43 - 1
MY_LINK_COLUMN = 6 - 1
ITEMPAGE_LINK_COLUMN = 112 - 1
update_PRODUCT_CODE_COLUMN = 2 - 1
TODAY = time.strftime("%Y%m%d")[-5:]
NOW = time.strftime("_%H%M")
SOLD_OUT_LINE = 2
HEADER_LIST = ["연번", "상품코드", "변경상품명", "상태", "옵션입력", "옵션개수", "OSO", "업데이트", "원본경로"]
DEFAULT_N_HEADER = len(HEADER_LIST)
initial_directory = "D:/Dropbox/A-TEB/작업폴더/999.품단종처리/"
# initial_directory = "C:/Users/JK/Dropbox/A-TEB/작업폴더/999.품단종처리/"
NO = 1
YES = 0
color_green = fg("light_green")
color_red = fg("red")
color_yellow = fg("yellow")
reset = attr("reset")


def set_Logger(name=None):
    # 1 logger instance를 만든다.
    logger = logging.getLogger(name)

    # 2 logger의 level을 가장 낮은 수준인 DEBUG로 설정해둔다.
    logger.setLevel(logging.INFO)

    # 3 formatter 지정
    # formatter = logging.Formatter("\n----- %(asctime)s / %(funcName)s / %(levelname)s : (%(lineno)d) %(message)s")
    formatter = logging.Formatter(
        ">> %(asctime)s / %(funcName)s / %(levelname)s : (%(lineno)d) %(message)s"
    )

    # 4 handler instance 생성
    console = logging.StreamHandler()

    log_filename = initial_directory + "log_" + TODAY + NOW + ".log"
    file_handler = logging.FileHandler(filename=log_filename)

    # 5 handler 별로 다른 level 설정
    console.setLevel(logging.INFO)
    file_handler.setLevel(logging.INFO)

    # 6 handler 출력 format 지정
    console.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # 7 logger에 handler 추가
    # logger.addHandler(console)
    logger.addHandler(file_handler)

    return logger


def open_ExcelFile(title):
    """
    엑셀파일 열기
    :return: 파일명 리턴
    """

    root = Tk()
    root.withdraw()
    root.filename = filedialog.askopenfilename(
        initialdir=initial_directory,
        title=title,
        filetypes=(("data files", "*.csv;*.xls;*.xlsx"), ("all files", "*.*")),
    )
    print("\nData Read From << ", root.filename)
    return root.filename


def wait_AnyInput():
    input("Enter Any Key to Proceed...")


def select_Mode():
    while True:

        print("\n\n======== 품단종 처리 ========")
        print("1. 신규 원장 생성")
        print("2. 기존 원장 업데이트")
        print("3. 품절/재입고 처리 - 전체 수집")
        print("4. 품절/재입고 처리 - 일부 수집 내용만 반영")
        print("5. 재수집 링크 생성")
        print("6. 셀러고 전처리 - 여러 엑셀모으기")
        print("7. 셀러고 전처리 - 한줄 만들기")
        print("8. 셀러고 - 옵션품절")
        print("9. 셀러고 - 판매재개")
        print("0. 종료")


        mode = int(input("\nSelect Mode : "))

        if mode == 1:
            return "CREATE_SB"
        elif mode == 2:
            return "UPDATE_SB"
        elif mode == 3:
            return "UPDATE_SO_T"
        elif mode == 4:
            return "UPDATE_SO_P"
        elif mode == 5:
            return "MAKE_LINK"
        elif mode == 6:
            return "GATHER_XLS"
        elif mode == 8:
            return "SLG_OSO"
        elif mode == 9:
            return "SLG_SA"
        elif mode == 0:
            return "EXIT"


def count_OptionNumber(option_info):
    """
    옵션 갯수 계산 후 리턴
    :param option_info: 옵션정보
    :return: 옵션 수
    """

    return option_info.count("**")


def count_MaxOptionNumber(option_list):
    """
    최대 옵션 갯수 계산후 리턴
    :param option_list: 옵션정보 컬럼 리스트
    :return: max 옵션 수
    """

    max_n_options = 0
    for option_info in option_list:
        n_options = count_OptionNumber(option_info)
        if n_options >= max_n_options:
            max_n_options = n_options
    return max_n_options


def remove_AddedChar(option):
    # return re.sub('\[(.*\)]|\s-\s.*','',re.sub('[.]+[0-9]+[\uac1c]+','',option)).split()[0]
    # return ' '.join(re.sub('[.]+[0-9]+[\uac1c]+','',option).split())
    # return ' '.join(re.sub("\[.*\]|\s-\s.*",'',re.sub('[.]+[0-9]+[\u3131-\u3163\uac00-\ud7a3]+','',option)).split())
    # return " ".join(
    #     re.sub(
    #         "\[.*\]|\s-\s.*",
    #         "",
    #         re.sub(
    #             r"\s+[.0-9]+[\u3131-\u3163\uac00-\ud7a3]{3}",
    #             "",
    #             option,
    #         ),
    #     ).split()
    # )
    # return re.sub("\[.*\]|\s-\s.*", "",
    #        re.sub("-{2,}", "", re.sub(r"\s{2,}[.,0-9]+[\u3131-\u3163\uac00-\ud7a3]{3}", "", option))).strip()
    # return re.sub("\[.*\]|\s-\s.*", "",re.sub("-{2,}", "", re.sub(r"[.,0-9]+[\u3131-\u3163\uac00-\ud7a3]{3}", "", option))).strip()
    return re.sub(
        "(?<=[가-힣]) +(?=[가-힣])",
        "",
        re.sub("\[.*\]|\s-\s.*",
               "",re.sub("-{2,}",
                         "",
                         re.sub(
                             r"[.,0-9]+[\u3131-\u3163\uac00-\ud7a3]{3}",
                             "",
                             option)
                         )
               ).strip()
    )

def make_OptionInfoData(type, option_cate, option_info):
    """
    옵션정보 받아서 옵션명 : 갯수 조합으로 바꿔 리턴
    :param option_cate: 옵션명
    :param option_info: 옵션정보
    옵션명에 따라 옵션정보의 옵션명 조합방식 변경 추가
    :return:
    """

    option_cate_list = option_cate.splitlines()
    option_list = option_info.splitlines()
    option_info_list = []

    for option in option_list:
        split_option_list = option.split("*")
        # split_option_list = option.replace(" ", "").split("*")
        # split_option_list = re.sub(r"(?<=[가-힣]) +(?=[가-힣])", "", option.split("*"))

        # 옵션명이 옵션정보에 들어있으면
        if (
            option_cate_list[0].replace(" ","") == split_option_list[0]
            or option_cate_list[1].replace(" ","") == split_option_list[0]
        ):
            option_name = remove_AddedChar(split_option_list[1].replace(",", ""))
        else:
            option_name = (
                remove_AddedChar(split_option_list[0].replace(",", ""))
                + " "
                + remove_AddedChar(split_option_list[1].replace(",", ""))
            )
        option_amount = int(split_option_list[5])
        option_info_list.append(option_name)
        option_info_list.append(option_amount)

        if type == "NEW":
            option_info_list.append("신규")
        elif type == "UPDATE":
            if int(option_amount) > SOLD_OUT_LINE:
                option_info_list.append("판매중")
            else:
                option_info_list.append("품절")
    return option_info_list


def excel_Visible():
    # from win32com.client import Dispatch
    excel = Dispatch('Excel.Application')
    excel.Visible = True
    excel.ScreenUpdating = True
    excel.DisplayAlerts = True
    excel.EnableEvents = True

def autofit_ColumnSize(mode, target_filename):
    excel = Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(target_filename)

    excel.Visible = False
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False

    """
    터미널 복구 코드 ; 엑셀 안보이는 경우 콘솔에서 복사 실행
    from win32com.client import Dispatch
    excel = Dispatch('Excel.Application')
    excel.Visible = True
    excel.ScreenUpdating = True
    excel.DisplayAlerts = True
    excel.EnableEvents = True
    """

    if mode == "O":
        excel.Worksheets(1).Activate()
        excel.ActiveSheet.Columns.AutoFit()
        excel.ActiveSheet.Columns(9).ColumnWidth = 2
    elif mode == "SO":
        for i in range(1, 5):
            excel.Worksheets(i).Activate()
            excel.ActiveSheet.Columns.AutoFit()
        excel.Worksheets(1).Activate()
    wb.Save()
    wb.Close()

    # excel_Visible()


def find_SourceCode(source_code, target_code_list):
    for index, target_code in enumerate(target_code_list):
        if source_code == target_code:
            return index
    return -1


def isIn_SourceOptionList(source_list, update_list):
    for update_option in update_list:
        for source_option in source_list:
            if update_option == source_option:
                return True
    return False


def open_SourceUpdateExcel(mode):
    """
    기준 원장과 업데이트 파일을 열어서 백업하고 파일명 리턴
    mode : UPDATE_SB, UPDATE_SO
    :return: source_filename, update_filename
    """

    # 원장엑셀 오픈
    print(reset)
    source_filename = open_ExcelFile("원장 엑셀파일 선택")  # 원장 선택

    print(f"=> 기준 원장 엑셀 파일 선택 : {source_filename}")
    logger.info(f"=> 기준 원장 엑셀 파일 선택 : {source_filename}")

    # 원장엑셀 백업
    backup_filename = source_filename[:-5] + NOW + "_BU.xlsx"
    shutil.copy2(source_filename, backup_filename)

    print(f"==> 백업파일 저장  : {backup_filename}")
    logger.info(f"==> 백업파일 저장  : {backup_filename}")

    # 수집엑셀 오픈
    if mode == "UPDATE_SB":
        update_filename = open_ExcelFile("업데이트할 누적데이터 엑셀파일 선택")  # 수집 데이터 선택
        print(f"=> 업데이트할 누적데이터 : {update_filename}")
        logger.info(f"=> 업데이트할 누적데이터 : {update_filename}")
    elif mode == "UPDATE_SO_T" or mode == "UPDATE_SO_P":
        update_filename = open_ExcelFile("업데이트할 수집 엑셀파일 선택")  # 수집 데이터 선택
        print(f"=> 업데이트할 신규수집 엑셀 : {update_filename}")
        logger.info(f"=> 업데이트할 신규수집 엑셀 : {update_filename}")

    return source_filename, update_filename

def create_SB():
    """
    원장 생성
    ateb 누적 데이터를 기준으로 기준원장 생성
    연번, 상품코드, 변경상품명, 상태코드, 재고, 판매량, 업데이트일자, 옵션갯수, 옵션명:재고...
    상태코드
    - 생성, 판매중, 품절, 옵션품절
    :return:
    """

    source_filename = open_ExcelFile("누적데이터 엑셀파일 선택")  # 누적데이터 오픈

    source_sheet = pyexcel.get_sheet(file_name=source_filename, name_columns_by_row=0)

    target_filename = source_filename[:-5] + "_" + TODAY + NOW + "_원장.xlsx"
    max_n_option = count_MaxOptionNumber(source_sheet.column["옵션정보"])

    target_sheet = pyexcel.Sheet()  # 새로운 엑셀 생성

    # 헤더
    option_header_list = []
    if max_n_option == 0:
        max_n_option = 1
    for option_header in range(1, 3 * max_n_option + 1, 3):
        option_header_list.append("옵션" + str(int(option_header / 3) + 1))
        option_header_list.append("재고" + str(int(option_header / 3) + 1))
        option_header_list.append("상태" + str(int(option_header / 3) + 1))
    target_sheet_header = HEADER_LIST + option_header_list

    for i in range(len(target_sheet_header)):
        target_sheet[0, i] = target_sheet_header[i]
    target_sheet.name_columns_by_row(0)

    # 원장 데이터 기록
    actual_data = 0
    passed_data = 0
    target_row_ptr = 0

    for row_ptr in range(source_sheet.number_of_rows()):

        source_product_code = source_sheet[row_ptr, "상품코드"]
        source_product_name = source_sheet[row_ptr, "변경상품명"]
        my_link = source_sheet[row_ptr, "마이링크"]
        if my_link == "" or my_link == None or my_link.find("_") == -1:
            passed_data += 1
            continue  # 작업안된 것은 패스
        n_option = count_OptionNumber(source_sheet[row_ptr, "옵션정보"])

        target_sheet[target_row_ptr, "연번"] = target_row_ptr
        target_sheet[target_row_ptr, "상품코드"] = source_product_code
        target_sheet[target_row_ptr, "변경상품명"] = source_product_name
        target_sheet[target_row_ptr, "상태"] = target_sheet[target_row_ptr, "상태1"] = "신규"
        target_sheet[target_row_ptr, "옵션입력"] = source_sheet[row_ptr, "옵션입력"]
        target_sheet[target_row_ptr, "옵션개수"] = n_option
        target_sheet[target_row_ptr, "업데이트"] = TODAY + NOW
        target_sheet[target_row_ptr, "원본경로"] = source_sheet[
            row_ptr, source_sheet.number_of_columns() - 1
        ]

        actual_data += 1

        if n_option == 0:
            target_sheet[target_row_ptr, "재고1"] = 999
            # target_sheet[target_row_ptr, 10] = 999
        else:
            # 옵션정보 리스트 생성
            option_list = make_OptionInfoData(
                "NEW", source_sheet[row_ptr, "옵션명"], source_sheet[row_ptr, "옵션정보"]
            )

            for col_ptr, option_value in enumerate(option_list):
                target_sheet[target_row_ptr, DEFAULT_N_HEADER + col_ptr] = option_value
        target_row_ptr += 1

    # 생성된 엑셀 저장
    print(f"  >> {passed_data} items passed from {source_sheet.number_of_rows()}")
    print(f"  >> Finally {actual_data} Data Saved to >> ", target_filename)

    target_sheet.save_as(target_filename)

    # 열넓이 조정
    autofit_ColumnSize("O", target_filename)

def update_SB(mode):
    """
    기존 원장에 없는 신규 누적데이터 추가

    기준 원장과 신규 누적데이터 열어서
    기준 원장에 없는 상품과 옵션 추가
    """

    # 데이터 오픈
    source_filename, update_filename = open_SourceUpdateExcel(mode)

    update_sheet = pyexcel.get_sheet(
        file_name=update_filename, name_columns_by_row=0
    )
    source_sheet = pyexcel.get_sheet(
        file_name=source_filename, name_columns_by_row=0
    )

    source_product_code_list = source_sheet.column["상품코드"]
    source_row_ptr = source_sheet.number_of_rows()
    n_not_processed_data = 0 # 작업전으로 패스한 데이터
    n_included_data = 0 # 소스에 있어서 패스한 데이터
    n_added_data = 0 # 추가된 데이터

    update_product_code_list = update_sheet.column["상품코드"]

    for update_row_ptr, update_product_code in enumerate(update_product_code_list):

        update_product_name = update_sheet[update_row_ptr, "변경상품명"]
        update_my_link = update_sheet[update_row_ptr, "마이링크"]
        n_option = count_OptionNumber(update_sheet[update_row_ptr, "옵션정보"])

        if update_my_link == "" or update_my_link == None or update_my_link.find("_") == -1:
            n_not_processed_data += 1
            continue  # 작업안된 것은 패스

        if update_product_code in source_product_code_list:
            # 상품코드가 기존 원장에 있으면
            n_included_data += 1
            continue
        else: # 상품코드가 기존 원장에 없으면 기존 원장 끝에 추가
            source_sheet[source_row_ptr, "연번"] = source_row_ptr
            source_sheet[source_row_ptr, "상품코드"] = update_product_code
            source_sheet[source_row_ptr, "변경상품명"] = update_product_name
            source_sheet[source_row_ptr, "상태"] = source_sheet[source_row_ptr, "상태1"] = "신규"
            source_sheet[source_row_ptr, "옵션입력"] = update_sheet[update_row_ptr, "옵션입력"]
            source_sheet[source_row_ptr, "옵션개수"] = n_option
            source_sheet[source_row_ptr, "업데이트"] = TODAY + NOW
            source_sheet[source_row_ptr, "원본경로"] = update_sheet[
                update_row_ptr, update_sheet.number_of_columns() - 1
            ]
            n_added_data += 1
            print(f"  >> {n_added_data} : {update_product_code} {update_product_name} added to row {source_row_ptr}.. ")

            if n_option == 0: # 단일상품 재고, 상태 처리
                source_sheet[source_row_ptr, "재고1"] = 999
                print(f"  >> {n_added_data} : => 단일상품 ")

            else: # 옵션있는 상품이면 옵션정보 리스트 생성
                option_list = make_OptionInfoData(
                    "NEW", update_sheet[update_row_ptr, "옵션명"], update_sheet[update_row_ptr, "옵션정보"]
                )

                for col_ptr, option_value in enumerate(option_list):
                    source_sheet[source_row_ptr, DEFAULT_N_HEADER + col_ptr] = option_value
                print(f"  >> {n_added_data} : => 옵션상품 {option_list}")
            source_row_ptr += 1

    # 생성된 엑셀 저장
    print(f" =>> {n_included_data} items already included.. ")
    print(f" =>> not finished {n_not_processed_data} items passed.. ")

    print(f" =>> Finally {n_added_data} Data Updated to >> ", source_filename)

    source_sheet.save_as(source_filename)



def update_SO(mode):
    """
    품절 확인
    기준원장과 상품수집엑셀 비교하여 품절 / 옵션품절 판정 및 기준원장 업데이트
    품절처리 대상, 판매재개 대상, 옵션품절처리 대상, 옵션판매재개 대상 판정
    :return:
    상태업데이트 ; 옵션갯수에 따라 별도 처리
    최종산물 : 품절처리 대상 목록, 판매재개 대상 목록, 옵션품절처리 대상 목록, 옵션판매재개 대상 목록
    """

    # 원장엑셀 오픈
    source_filename, update_filename = open_SourceUpdateExcel(mode)

    # 데이터 리딩 준비
    update_sheet = pyexcel.get_sheet(file_name=update_filename, name_columns_by_row=0)
    source_sheet = pyexcel.get_sheet(file_name=source_filename, name_columns_by_row=0)

    source_product_code_list = source_sheet.column["상품코드"]
    source_product_name_list = source_sheet.column["변경상품명"]
    source_product_link_list = source_sheet.column["원본경로"]

    update_product_code_list = update_sheet.column["상품코드"]

    # 결과 엑셀파일 준비

    SO_filename = source_filename[:-18] + TODAY + NOW + "_SO.xlsx"  # 품절, 판매재개 파일

    SO_sheet = pyexcel.Sheet()
    SO_sheet[0,0] = ""
    SO_sheet[0,1] = ""
    SA_sheet = pyexcel.Sheet()
    OSO_sheet = pyexcel.Sheet()
    OSA_sheet = pyexcel.Sheet()

    SO_row_ptr = 0  # 결과 엑셀 행포인터
    SA_row_ptr = 0

    # 판정 루프 ; 소스 상품별 루프

    OSO_list = []
    OSO_link_list = []
    OSA_list = []
    OSA_link_list = []

    for source_code_row_index, source_product_code in enumerate(
        source_product_code_list
    ):

        current_source_option_list = []
        current_source_stock_list = []
        current_source_status_list = []

        current_update_option_list = []
        current_update_stock_list = []
        current_update_status_list = []

        # 회차
        print(
            f"\n>> Process {source_code_row_index + 1} / {len(source_product_code_list)} : {source_product_code}"
        )
        logger.info(
            f">> Process {source_code_row_index + 1} / {len(source_product_code_list)} : {source_product_code}"
        )

        # 전처리2: 수집엑셀의 행번호 판정 ; 업데이트 목록에 포함여부
        # 전처리2: 소스에만 있고 업데이트에 없는 상품은 품절처리; 단종된 것임.

        PRE1_ONLYIN_SOURCE = 0

        if mode == "UPDATE_SO_P":
            print(color_red + "  > 전처리 1 =>> 실행모드에 의해 미수행")
            logger.info("  > 전처리 1 =>> 실행모드에 의해 미수행")
        else:
            print(color_red + "  > 전처리 1 : 업데이트 목록 포함여부 확인")
            logger.info("  > 전처리 1 : 업데이트 목록 포함여부 확인")

        # 업데이트 포함여부 판정
        PRE1_ONLYIN_SOURCE = 0
        update_code_row_index = find_SourceCode(
            source_product_code, update_product_code_list
        )
        source_n_option = source_sheet[source_code_row_index, "옵션개수"]

        if update_code_row_index != -1:  # 업데이트파일에 포함되어 있으면
            print(
                f'      {source_code_row_index} : {source_product_code}  ==> {update_code_row_index} : {update_sheet[update_code_row_index, "상품코드"]} '
            )
            logger.info(
                f'      {source_code_row_index} : {source_product_code}  ==> {update_code_row_index} : {update_sheet[update_code_row_index, "상품코드"]} '
            )

            source_n_option = int(
                source_sheet[source_code_row_index, "옵션개수"]
            )  # 소스 옵션개수
            update_n_option = int(
                update_sheet[update_code_row_index, "옵션개수"]
            )  # 업데이트 옵션개수

            if source_n_option > 0 or update_n_option > 0:  # 옵션상품이면 옵션리스트 생성

                for i in range(source_n_option):
                    current_source_option_list.append(
                        source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]
                    )
                    current_source_stock_list.append(
                        source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3+1]
                    )
                    current_source_status_list.append(
                        source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3+2]
                    )
                current_update_option_list = make_OptionInfoData(
                    "UPDATE",
                    update_sheet[update_code_row_index, "옵션명"],
                    update_sheet[update_code_row_index, "옵션정보"],
                )[::3]
                current_update_stock_list = make_OptionInfoData(
                    "UPDATE",
                    update_sheet[update_code_row_index, "옵션명"],
                    update_sheet[update_code_row_index, "옵션정보"],
                )[1::3]
                current_update_status_list = make_OptionInfoData(
                    "UPDATE",
                    update_sheet[update_code_row_index, "옵션명"],
                    update_sheet[update_code_row_index, "옵션정보"],
                )[2::3]
            update_stock = int(
                update_sheet[update_code_row_index, "재고수량"]
            )  # 업데이트 재고수량
        elif mode == "UPDATE_SO_T" and update_code_row_index == -1:  # 모드 T + 업데이트 목록에 없으면 품절처리

            PRE1_ONLYIN_SOURCE = 1
            update_stock = 0
            print(
                f"{source_code_row_index} : {source_product_code}  ==> {update_code_row_index} : 업데이트 목록에 없음 "
            )
            logger.info(
                f"{source_code_row_index} : {source_product_code}  ==> {update_code_row_index} : 업데이트 목록에 없음 "
            )

            current_update_option_list = []
            current_update_stock_list = []
            current_update_status_list = []

            # PRE1_ONLYIN_SOURCE = 1

            print("  > 전처리 1 => 수집 누락 있음")
            logger.info("  > 전처리 1 => 수집 누락 있음")
            # 품절아니면 품절로 상태 변경
            print(
                f"  > 전처리 1 => 수집 누락 있음 >> {source_product_code} {source_product_name_list[source_code_row_index]} 현재 상태 {source_sheet[source_code_row_index, '상태']}"
            )
            logger.info(
                f"  > 전처리 1 => 수집 누락 있음 >> {source_product_code} {source_product_name_list[source_code_row_index]} 현재 상태 {source_sheet[source_code_row_index, '상태']}"
            )
            if source_sheet[source_code_row_index, "상태"] != "품절":
                source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                source_sheet[source_code_row_index, "상태"] = "전처리품절"
                # 품절처리목록에 추가
                SO_sheet[SO_row_ptr, 0] = source_product_code
                SO_sheet[SO_row_ptr, 1] = source_product_name_list[
                    source_code_row_index
                ]
                SO_sheet[SO_row_ptr, 2] = source_product_link_list[
                    source_code_row_index
                ]
                SO_row_ptr += 1
                print(
                    f"  > 전처리 1 => >> SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                )
                logger.info(
                    f"  > 전처리 1 => >> SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                )

                if source_n_option == 0:  # 단일상품 재고변경
                    source_sheet[source_code_row_index, "재고1"] = 0
                    source_sheet[source_code_row_index, "상태1"] = "전처리품절"
                    print(f"  > 전처리 1 => >> 단일상품 재고변경; 전처리 품절처리")
                    logger.info(f"  > 전처리 1 => >> 단일상품 재고변경; 전처리 품절처리")
                else:
                    # 옵션상품 품절처리 및 재고 변경
                    source_sheet[source_code_row_index, "OSO"] = source_n_option
                    print(f"  > 전처리 1 => >> 옵션상품 재고변경; 전체옵션 전처리 품절처리")
                    logger.info(f"  > 전처리 1 => >> 옵션상품 재고변경; 전체옵션 전처리 품절처리")
                    for i in range(source_n_option):
                        if (
                            source_sheet[
                                source_code_row_index, DEFAULT_N_HEADER + i * 3 + 2
                            ]
                            != "품절"
                        ):
                            source_sheet[
                                source_code_row_index, DEFAULT_N_HEADER + i * 3 + 1
                            ] = 0
                            source_sheet[
                                source_code_row_index, DEFAULT_N_HEADER + i * 3 + 2
                            ] = "전처리품절"
                            OSO_list.append(
                                source_sheet[source_code_row_index, "상품코드"]
                                + " "
                                + source_sheet[
                                    source_code_row_index, DEFAULT_N_HEADER + i * 3
                                ]
                            )
                            OSO_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                            print(
                                f"  > 전처리 1 => >> OSO_list에 추가 {source_product_code} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                            )
                            logger.info(
                                f"  > 전처리 1 => >> OSO_list에 추가 {source_product_code} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                            )
            elif source_sheet[source_code_row_index, "상태"] == "품절":
                print(f"  > 전처리 1 => >> 업데이트 누락 + 기존 상태 품절로 PASS")
                logger.info(f"  > 전처리 1 => >> 업데이트 누락 + 기존 상태 품절로 PASS")
                continue
        else:
            print("  > 전처리 1 => 수집 누락 없음")
            logger.info("  > 전처리 1 => 수집 누락 없음")

        # 전처리2: 옵션입력 바뀌면 옵션입력변경 리스트에 추가
        PRE2_OPTION_TYPE_CHANGED = 0
        if PRE1_ONLYIN_SOURCE != 1:
            print(color_green + "  > 전처리 2 : 옵션입력 변경 확인")
            logger.info("  > 전처리 2 : 옵션입력 변경 확인")

            if source_sheet[source_code_row_index, "옵션입력"] == "조합형" and (
                update_sheet[update_code_row_index, "옵션입력"] == ""
                or update_sheet[update_code_row_index, "옵션입력"] == None
            ):

                print("  > 전처리 2 => 옵션입력 변경으로 단종 품절 처리")
                logger.info("  > 전처리 2 => 옵션입력 변경으로 단종 품절 처리")

                PRE2_OPTION_TYPE_CHANGED = 1

                # 품절아니면 품절로 상태 변경
                if source_sheet[source_code_row_index, "상태"] != "품절":
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[source_code_row_index, "상태"] = "전처리품절"

                    # 품절처리목록에 추가
                    SO_sheet[SO_row_ptr, 0] = source_product_code
                    SO_sheet[SO_row_ptr, 1] = source_product_name_list[
                        source_code_row_index
                    ]
                    SO_sheet[SO_row_ptr, 2] = source_product_link_list[
                        source_code_row_index
                    ]
                    SO_row_ptr += 1
                    print(
                        f"  > 전처리 2 => >> SO_sheet에 {SO_row_ptr-1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )
                    logger.info(
                        f"  > 전처리 2 => >> SO_sheet에 {SO_row_ptr-1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )

                    if source_n_option == 0:  # 단일상품 재고변경
                        source_sheet[source_code_row_index, "재고1"] = 0
                        source_sheet[source_code_row_index, "상태1"] = "전처리품절"
                        print(f"  > 전처리 1 => >> 단일상품 재고변경; 전처리 품절처리")
                        logger.info(f"  > 전처리 1 => >> 단일상품 재고변경; 전처리 품절처리")
                    else:
                        # 옵션상품 품절처리 및 재고 변경
                        print(f"  > 전처리 1 => >> 옵션상품 재고변경; 전체옵션 전처리 품절처리")
                        logger.info(f"  > 전처리 1 => >> 옵션상품 재고변경; 전체옵션 전처리 품절처리")

                        source_sheet[source_code_row_index, "OSO"] = source_n_option

                        for i in range(source_n_option):
                            source_sheet[
                                source_code_row_index, DEFAULT_N_HEADER + i * 3 + 1
                            ] = 0
                            source_sheet[
                                source_code_row_index, DEFAULT_N_HEADER + i * 3 + 2
                            ] = "전처리품절"
                            OSO_list.append(
                                source_sheet[source_code_row_index, "상품코드"]
                                + " "
                                + source_sheet[
                                    source_code_row_index, DEFAULT_N_HEADER + i * 3
                                ]
                            )
                            OSO_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                            print(
                                f"  > 전처리 1 => >> OSO_list에 추가 {source_product_code} {source_product_name_list[source_code_row_index]} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                            )
                            logger.info(
                                f"  > 전처리 1 => >> OSO_list에 추가 {source_product_code} {source_product_name_list[source_code_row_index]} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                            )
                # continue
            else:
                print("  > 전처리 2 => 옵션입력 변경 없음")
                logger.info("  > 전처리 2 => 옵션입력 변경 없음")
        else:
            print(color_red + "  > 전처리 2 => 전처리 1 결과에 의해 PASS")
            logger.info("  > 전처리 2 => 전처리 1 결과에 의해 PASS")

        # 전처리3: 원장에 없고 업데이트에만 있는 옵션은 원장에 추가

        if PRE2_OPTION_TYPE_CHANGED != 1 and PRE1_ONLYIN_SOURCE != 1:

            print(color_yellow + "  > 전처리 3 : 업데이트 목록에만 있는 신규 옵션은 원장에 추가")
            logger.info("  > 전처리 3 : 업데이트 목록에만 있는 신규 옵션은 원장에 추가")

            SA_ADDED = 0
            LIST_UPDATED = 0

            # 현재 업데이트 옵션이 원장에 없으면
            for update_option_index, current_update_option in enumerate(
                current_update_option_list
            ):
                if current_update_option not in current_source_option_list:
                    # 원장에 업데이트 옵션 추가
                    print(
                        f"  > 전처리 3 : 현재 업데이트 옵션 {current_update_option} => 현재 소스 옵션 목록 {current_source_option_list}에 없음"
                    )
                    logger.info(
                        f"  > 전처리 3 : 현재 업데이트 옵션 {current_update_option} => 현재 소스 옵션 목록 {current_source_option_list}에 없음"
                    )

                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW

                    current_update_stock = current_update_stock_list[
                        update_option_index
                    ]
                    source_sheet[
                        source_code_row_index, DEFAULT_N_HEADER + source_n_option * 3
                    ] = current_update_option
                    print(f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option}")
                    logger.info(f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option}")

                    if current_update_stock > SOLD_OUT_LINE:  # 재고 있으면 판매중 처리
                        source_sheet[
                            source_code_row_index,
                            DEFAULT_N_HEADER + source_n_option * 3 + 1,
                        ] = current_update_stock
                        source_sheet[
                            source_code_row_index,
                            DEFAULT_N_HEADER + source_n_option * 3 + 2,
                        ] = "판매중"
                        print(
                            f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option} >> 재고있음 >> 판매중 처리"
                        )
                        logger.info(
                            f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option} >> 재고있음 >> 판매중 처리"
                        )

                        # 추가 수집 리스트에 추가
                        if SA_ADDED == 0:
                            SA_sheet[SA_row_ptr, 0] = source_product_code
                            SA_sheet[SA_row_ptr, 1] = source_product_name_list[
                                source_code_row_index
                            ]
                            SA_sheet[SA_row_ptr, 2] = source_product_link_list[
                                source_code_row_index
                            ]
                            SA_row_ptr += 1
                            print(
                                f"  > 전처리 3    >> SA_sheet에 추가 {SA_row_ptr-1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                            )
                            logger.info(
                                f"  > 전처리 3    >> SA_sheet에 추가 {SA_row_ptr-1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                            )
                            SA_ADDED = 1
                    else:  # 재고 없으면 품절처리
                        print(
                            f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option} >> 재고없음 >> 품절 처리"
                        )
                        logger.info(
                            f"  > 전처리 3 ==> 원장추가 필요 >> {current_update_option} >> 재고없음 >> 품절 처리"
                        )
                        source_sheet[
                            source_code_row_index,
                            DEFAULT_N_HEADER + source_n_option * 3 + 1,
                        ] = 0
                        source_sheet[
                            source_code_row_index,
                            DEFAULT_N_HEADER + source_n_option * 3 + 2,
                        ] = "전처리품절"
                        OSO_list.append(
                            source_sheet[source_code_row_index, "상품코드"]
                            + " "
                            + current_update_option
                        )
                        OSO_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                        print(
                            f"  > 전처리 3    >> OSO_list에 추가 {source_product_code} {current_update_option}"
                        )
                        logger.info(
                            f"  > 전처리 3    >> OSO_list에 추가 {source_product_code} {current_update_option}"
                        )
                    source_n_option += 1
                    print(f"  > 전처리 3    >> 총 옵션개수 업데이트 : {source_n_option}")
                    logger.info(f"  > 전처리 3    >> 총 옵션개수 업데이트 : {source_n_option}")
                    source_sheet[source_code_row_index, "옵션개수"] = source_n_option
                    LIST_UPDATED = 1

            if LIST_UPDATED == 0:
                print("  > 전처리 3  ==> 없음")
                logger.info("  > 전처리 3 ==> 없음")
        else:
            print(color_red + "  > 전처리 3 => 전처리 1, 2 결과에 의해 PASS")
            logger.info("  > 전처리 3 => 전처리 1, 2 결과에 의해 PASS")

        # 메인: 소스와 업데이트에 모두 있는 옵션은 업데이트하고 품절/판매중처리

        if mode == "UPDATE_SO_P" and find_SourceCode(source_product_code, update_product_code_list) == -1: # 모드 T + 소스에 없으면
            print(color_green + "  > 메인 =>> 실행모드에 의거 미수행")
            logger.info("  > 메인 =>> 실행모드에 의거 미수행")
            continue

        print(color_green + "  > 메인 : 원장 분석")
        logger.info("  > 메인 : 원장 분석")

        source_status = source_sheet[source_code_row_index, "상태"]

        # 1. 단일상품인데
        if source_n_option == 0:

            print("  > 메인 => 단일상품")
            logger.info("  > 메인 => 단일상품")

            # 업데이트 재고있고

            if update_stock > SOLD_OUT_LINE:
                print("  > 메인 => 단일상품 --> 재고있음")
                logger.info("  > 메인 => 단일상품 --> 재고있음")

                if source_status == "전처리품절":
                    # 재고 있어보여도 전처리품절이면 무조건 품절상태로 변경 ; 품절목록에는 전처리 과정에서 추가
                    source_sheet[source_code_row_index, "상태"] = "품절"
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[source_code_row_index, "재고1"] = 0

                    print("  > 메인 => 단일상품  --> 전처리품절 >> 품절 변경")
                    logger.info("  > 메인 => 단일상품  --> 전처리품절 >> 품절 변경")
                elif source_status == "품절" or source_status == "신규":
                    if source_status == "품절":
                        # 품절이었으면 재개목록에 추가하고 업데이트
                        SA_sheet[SA_row_ptr, 0] = source_product_code
                        SA_sheet[SA_row_ptr, 1] = source_product_name_list[
                            source_code_row_index
                        ]
                        SA_sheet[SA_row_ptr, 2] = source_product_link_list[
                            source_code_row_index
                        ]
                        SA_row_ptr += 1
                        print(
                            f"  메인 => 단일상품 => 원장품절에서 재입고로 판매재개목록에 추가 >> SA_sheet에 추가 {SA_row_ptr-1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                        )
                        logger.info(
                            f"  메인 => 단일상품 => 원장품절에서 재입고로 판매재개목록에 추가 >> SA_sheet에 추가 {SA_row_ptr-1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                        )
                    elif source_status == "신규":
                        # 신규면 판매중으로
                        print("  > 메인 => 단일상품 --> 재고있음 --> 신규 >> 판매중")
                        logger.info("  > 메인 => 단일상품 --> 재고있음 --> 신규 >> 판매중")
                    source_sheet[source_code_row_index, "상태"] = "판매중"
                    source_sheet[source_code_row_index, "상태1"] = "판매중"
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[source_code_row_index, "재고1"] = update_stock
                elif source_status == "판매중":
                    # 이미 판매중이었다면 재고수량 반영만
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[source_code_row_index, "상태1"] = "판매중"
                    source_sheet[source_code_row_index, "재고1"] = update_stock
                    print("  > 메인 => 단일상품 --> 재고있음 --> 판매중 유지")
                    logger.info("  > 메인 => 단일상품 --> 재고있음 --> 판매중 유지")
            # 재고없고
            elif update_stock <= SOLD_OUT_LINE:
                print("  > 메인 => 단일상품 --> 재고없음")
                logger.info("  > 메인 => 단일상품 --> 재고없음")
                # 신규 또는 판매중이면
                if source_status == "신규" or source_status == "판매중":
                    # 품절로
                    source_sheet[source_code_row_index, "상태"] = "품절"
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[source_code_row_index, "재고1"] = update_stock
                    # 품절 리스트에 추가
                    print("  > 메인 => 단일상품 --> 재고없음 --> 신규/판매중 >> 품절")
                    logger.info("  > 메인 => 단일상품 --> 재고없음 --> 신규/판매중 >> 품절")
                    SO_sheet[SO_row_ptr, 0] = source_product_code
                    SO_sheet[SO_row_ptr, 1] = source_product_name_list[
                        source_code_row_index
                    ]
                    SO_sheet[SO_row_ptr, 2] = source_product_link_list[
                        source_code_row_index
                    ]
                    SO_row_ptr += 1
                    print(
                        f"  메인 => 단일상품 >> SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )
                    logger.info(
                        f"  메인 => 단일상품 >> SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )
                # 이미 품절이었다면 업데이트 기록만
                elif source_status == "품절" or source_status == "전처리품절":
                    if source_status == "전처리품절":
                        source_sheet[source_code_row_index, "상태"] = "품절"
                        print("  > 메인 => 단일상품 --> 재고없음 --> 전처리품절 >> 품절 변경")
                        logger.info("  > 메인 => 단일상품 --> 재고없음 --> 전처리품절 >> 품절 변경")
                    else:
                        print("  > 메인 => 단일상품 --> 재고없음 --> 품절 유지")
                        logger.info("  > 메인 => 단일상품 --> 재고없음 --> 품절 유지")
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
        # 2. 옵션있는 상품인 경우
        elif source_n_option >= 1:

            print(color_red + "  > 메인 => 옵션상품")
            logger.info("  > 메인 => 옵션상품")

            # 전처리 2-1. 소스에는 있는데 업데이트에는 없는 옵션은 품절처리

            # # 옵션 정보 리스트로 변환
            current_source_option_list = []
            current_source_stock_list = []
            current_source_status_list = []
            for i in range(source_n_option):
                current_source_option_list.append(
                    source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]
                )
                current_source_stock_list.append(
                    source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3 + 1]
                )
                current_source_status_list.append(
                    source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3 + 2]
                )

            print(color_red + "  > 메인 => 옵션상품 => 옵션 분석")
            logger.info("  > 메인 => 옵션상품 => 옵션 분석")

            OSO_count = 0  # 옵션품절 수 체크용

            print(f"  > 메인 => 옵션상품 >> 소스 옵션 목록 : {current_source_option_list}")
            print(f"  > 메인 => 옵션상품   >> 소스 재고 목록 : {current_source_stock_list}")
            print(f"  > 메인 => 옵션상품   >> 소스 상태 목록 : {current_source_status_list}")

            print(f"  > 메인 => 옵션상품 >> 업데이트 옵션 목록 : {current_update_option_list}")
            print(f"  > 메인 => 옵션상품   >> 업데이트 재고 목록 : {current_update_stock_list}")
            print(f"  > 메인 => 옵션상품   >> 업데이트 상태 목록 : {current_update_status_list}")

            logger.info(f"  > 메인 => 옵션상품 >> 소스 옵션 목록 : {current_source_option_list}")
            logger.info(f"  > 메인 => 옵션상품   >> 소스 재고 목록 : {current_source_stock_list}")
            logger.info(f"  > 메인 => 옵션상품   >> 소스 상태 목록 : {current_source_status_list}")

            logger.info(f"  > 메인 => 옵션상품 >> 업데이트 옵션 목록 : {current_update_option_list}")
            logger.info(f"  > 메인 => 옵션상품   >> 업데이트 재고 목록 : {current_update_stock_list}")
            logger.info(f"  > 메인 => 옵션상품   >> 업데이트 상태 목록 : {current_update_status_list}")

            # 기존, 업데이트 옵션정보 비교 판정 루프
            # 각 행별로 소스 옵션개수 만큼 루프

            SA_ADDED = 0
            for current_source_option_index, current_source_option in enumerate(current_source_option_list):
                print(f"  > 메인 => 옵션상품 >> 체크 대상 소스 옵션(현재) : {current_source_option}")
                # 소스 옵션개수만큼 루프, 기존 정보 체크
                # 소스는 전처리된 상태로 추가될 옵션이 있으면 이미 추가되어 있는 상태

                current_source_option_status = source_sheet[
                    source_code_row_index,
                    DEFAULT_N_HEADER + current_source_option_index * 3 + 2,
                ]
                print(
                    f"  > 메인 => 옵션상품 >> 체크 대상 소스 옵션 상태 (현재) : {current_source_option_status}"
                )
                logger.info(
                    f"  > 메인 => 옵션상품 >> 체크 대상 소스 옵션 상태 (현재) : {current_source_option_status}"
                )

                # 전처리품절이면 품절로 업데이트하고 다음 옵션으로
                if current_source_option_status == "전처리품절":
                    print(
                        f"  **> 메인 => 옵션상품 >> 체크 대상 소스 옵션 상태 (현재) : {current_source_option_status} ->> 전처리품절 >> 품절로 업데이트 후 PASS"
                    )
                    OSO_count += 1  # 옵션품절이니 OSO count 1 증가
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[
                        source_code_row_index,
                        DEFAULT_N_HEADER + current_source_option_index * 3 + 2,
                    ] = "품절"
                    continue

                # 현재 옵션이 업데이트옵션목록에 없으면 품절처리하고 다음옵션으로 ; 윗줄 전처리때문에 실제로는 실행되지 않음.
                if current_source_option not in current_update_option_list:
                    OSO_list.append(
                        source_sheet[source_code_row_index, "상품코드"]
                        + " "
                        + current_source_option
                    )
                    OSO_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                    OSO_count += 1
                    source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                    source_sheet[
                        source_code_row_index,
                        DEFAULT_N_HEADER + current_source_option_index * 3 + 1,
                    ] = 0
                    source_sheet[
                        source_code_row_index,
                        DEFAULT_N_HEADER + current_source_option_index * 3 + 2,
                    ] = "품절"
                    print(
                        f"  **> 메인 => 옵션상품 >> 소스옵션이 업데이트목록에 없어 품절처리 >> OSO_list에 추가 OSO {OSO_count} : {current_source_option}"
                    )
                    logger.info(
                        f" **> 메인 => 옵션상품 >> 소스옵션이 업데이트목록에 없어 품절처리 >> OSO_list에 추가 OSO {OSO_count} : {current_source_option}"
                    )
                    continue
                # 소ㅅ, 업데이트 옵션 비교
                for current_update_option_index, current_update_option in enumerate(
                    current_update_option_list
                ):
                    current_update_stock = current_update_stock_list[
                        current_update_option_index
                    ]
                    print(
                        f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} : {current_update_option}"
                    )
                    logger.info(
                        f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} : {current_update_option}"
                    )

                    # 원장에서 업데이트 대상 옵션 찾아서
                    if current_source_option == current_update_option:
                        print(
                            f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} = {current_update_option}"
                        )
                        logger.info(
                            f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} = {current_update_option}"
                        )

                        # 재고 있으면
                        if current_update_stock > SOLD_OUT_LINE:
                            print(
                                f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} 재고있음 >> 옵션 판매중 처리"
                            )
                            logger.info(
                                f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} 재고있음 >> 옵션 판매중 처리"
                            )
                            source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                            # 품절->판매중 변경시 판매재개로
                            if current_source_option_status == "품절":

                                OSA_list.append(
                                    source_sheet[source_code_row_index, "상품코드"]
                                    + " "
                                    + current_source_option
                                )
                                OSA_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                                print(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 품절 -> 판매중 처리 >> OSA_list에 추가 : {source_product_code} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                                )
                                logger.info(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 품절 -> 판매중 처리 >> OSA_list에 추가 : {source_product_code} {source_sheet[source_code_row_index, DEFAULT_N_HEADER + i * 3]}"
                                )
                                # 상품도 추가 수집 리스트에 추가; 불필요 할 수도...
                                if SA_ADDED == 0:
                                    SA_sheet[SA_row_ptr, 0] = source_product_code
                                    SA_sheet[
                                        SA_row_ptr, 1
                                    ] = source_product_name_list[source_code_row_index]
                                    SA_sheet[
                                        SA_row_ptr, 2
                                    ] = source_product_link_list[source_code_row_index]
                                    SA_row_ptr += 1
                                    print(
                                        f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 품절 -> 판매중 처리 >> SA_sheet에 추가 {SA_row_ptr - 1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                                    )
                                    logger.info(
                                        f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 품절 -> 판매중 처리 >> SA_sheet에 추가 {SA_row_ptr - 1} -> {SA_row_ptr} : {source_product_code} {source_product_name_list[source_code_row_index]}"
                                    )
                                    SA_ADDED = 1
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 1,
                                ] = current_update_stock
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 2,
                                ] = "판매중"
                                break
                            elif current_source_option_status != "품절":
                                print(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 상태 유지"
                                )
                                logger.info(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 상태 유지"
                                )
                                source_sheet[source_code_row_index, "업데이트"] = (
                                    TODAY + NOW
                                )
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 1,
                                ] = current_update_stock
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 2,
                                ] = "판매중"
                        # 재고 없으면
                        elif current_update_stock <= SOLD_OUT_LINE:
                            print(
                                f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} 재고없음 >> 옵션 품절 처리"
                            )
                            logger.info(
                                f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} 재고없음 >> 옵션 품절 처리"
                            )
                            source_sheet[source_code_row_index, "업데이트"] = TODAY + NOW
                            # 품절 아닌것->품절은 옵션품절로
                            if current_source_option_status != "품절":
                                OSO_count += 1
                                OSO_list.append(
                                    source_sheet[source_code_row_index, "상품코드"]
                                    + " "
                                    + current_update_option
                                )
                                OSO_link_list.append(source_sheet[source_code_row_index, "원본경로"])
                                print(
                                    f'  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 신규/판매중 -> 옵션 품절 처리 >> OSO_list에 추가 : OSO {OSO_count} {source_sheet[source_code_row_index, "상품코드"] + " " + current_update_option}'
                                )
                                logger.info(
                                    f'  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option} >> 신규/판매중 -> 옵션 품절 처리 >> OSO_list에 추가 : OSO {OSO_count} {source_sheet[source_code_row_index, "상품코드"] + " " + current_update_option}'
                                )
                            source_sheet[
                                source_code_row_index,
                                DEFAULT_N_HEADER + current_source_option_index * 3 + 1,
                            ] = 0
                            source_sheet[
                                source_code_row_index,
                                DEFAULT_N_HEADER + current_source_option_index * 3 + 2,
                            ] = "품절"
                            # 품절은 업데이트표시만
                            if (
                                current_source_option_status == "품절"
                                or current_source_option_status == "전처리품절"
                            ):
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 1,
                                ] = 0
                                source_sheet[
                                    source_code_row_index,
                                    DEFAULT_N_HEADER
                                    + current_source_option_index * 3
                                    + 2,
                                ] = "품절"
                                OSO_count += 1
                                print(
                                    f"   ----> 옵션품절 >> 옵션품절 처리목록에 미추가, OSO : {OSO_count} "
                                )
                                logger.info(
                                    f"   ----> 옵션품절 >> 옵션품절 처리목록에 미추가, OSO : {OSO_count} "
                                )
                                print(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option}  >> 옵션 품절 유지 OSO count만 추가 : OSO {OSO_count} {current_update_option}"
                                )
                                logger.info(
                                    f"  > 메인 => 옵션상품 >> 옵션 비교중 >> {current_source_option}  >> 옵션 품절 유지 OSO count만 추가 : OSO {OSO_count} {current_update_option}"
                                )
                            break
            source_sheet[source_code_row_index, "OSO"] = OSO_count
            print(
                f'   ==> 옵션상품 {source_sheet[source_code_row_index, "상태"]} : 옵션 {source_n_option}개 중 {OSO_count}개 품절'
            )
            logger.info(
                f'   ==> 옵션상품 {source_sheet[source_code_row_index, "상태"]} : 옵션 {source_n_option}개 중 {OSO_count}개 품절'
            )
            if OSO_count == source_n_option:
                source_sheet[source_code_row_index, "상태"] = "품절"
                print(f"   ----> >> 전체옵션 품절 처리")
                logger.info(f"   ----> >> 전체옵션 품절 처리")
                #SO list에 없으면 SO에 추가
                if len(SO_sheet.column[0]) == 0 or source_sheet[source_code_row_index, "상품코드"] not in SO_sheet.column[0]:
                    SO_sheet[SO_row_ptr, 0] = source_product_code
                    SO_sheet[SO_row_ptr, 1] = source_product_name_list[
                        source_code_row_index
                    ]
                    SO_sheet[SO_row_ptr, 2] = source_product_link_list[
                        source_code_row_index
                    ]
                    SO_row_ptr += 1
                    print(
                        f"   ----> >> 전체옵션 품절 처리 => SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )
                    logger.info(
                        f"   ----> >> 전체옵션 품절 처리 => SO_sheet에 {SO_row_ptr - 1} -> {SO_row_ptr} : {source_product_code}, {source_product_name_list[source_code_row_index]} 추가"
                    )
                #OSO list에서는 OSO 제거
                idx = 0
                while idx < len(OSO_list):
                    soldout_option = OSO_list[idx]
                    if soldout_option.split()[0] == source_product_code:
                        print(f"   ----> >> 전체옵션 품절 처리 => OSO_list에서  {OSO_list[idx]} 제거")
                        logger.info(f"   ----> >> 전체옵션 품절 처리 => OSO_list에서  {OSO_list[idx]} 제거")
                        del OSO_link_list[OSO_list.index(soldout_option)]
                        OSO_list.remove(soldout_option)
                        continue
                    idx += 1

            elif OSO_count == 0:
                source_sheet[source_code_row_index, "상태"] = "판매중"
                print(f"   ----> >> 전체옵션 판매중 처리")
                logger.info(f"   ----> >> 전체옵션 판매중 처리")
            else:
                source_sheet[source_code_row_index, "상태"] = "옵션품절"
                print(f"   ----> >> 전체옵션 옵션품절 처리")
                logger.info(f"   ----> >> 전체옵션 옵션품절 처리")
            print(reset)

    for row_ptr, OSO in enumerate(OSO_list):
        OSO_sheet[row_ptr, 0] = OSO
        OSO_sheet[row_ptr, 1] = OSO_link_list[row_ptr]

    if len(OSA_list) != 0:
        for row_ptr, OSA in enumerate(OSA_list):
            OSA_sheet[row_ptr, 0] = OSA
            OSA_sheet[row_ptr, 1] = OSA_link_list[row_ptr]

    SO_sheet.name = "SO"
    SA_sheet.name = "SA"

    OSO_sheet.name = "OSO"
    OSA_sheet.name = "OSA"

    SO_book = SO_sheet + SA_sheet + OSO_sheet + OSA_sheet

    # 생성된 엑셀 저장
    print(reset)
    print("\nOriginal Data Saved to >>", source_filename)
    logger.info(f"\nOriginal Data Saved to >> {source_filename}")
    print(
        f"  Total {len(source_product_code_list)} items / {len(update_product_code_list)} items checked.."
    )
    logger.info(
        f"  Total {len(source_product_code_list)} items / {len(update_product_code_list)} items checked.."
    )
    source_sheet.save_as(source_filename)

    print("\nProcessed Data Saved to >>", SO_filename)
    logger.info(f"\nProcessed Data Saved to >> {SO_filename}")
    print(f"  {SO_row_ptr} items SOLD OUT.. {SA_row_ptr} items Restocked")
    logger.info(f"  {SO_row_ptr} items SOLD OUT.. {SA_row_ptr} items Restocked")
    SO_book.save_as(SO_filename)
    SO_backup_filename = SO_filename[:-5] + NOW + "_BU.xlsx"
    shutil.copy2(SO_filename, SO_backup_filename)

    print(
        f"  {len(OSO_list)} option items SOLD OUT.. {len(OSA_list)} option items Restocked.."
    )
    logger.info(
        f"  {len(OSO_list)} option items SOLD OUT.. {len(OSA_list)} option items Restocked.."
    )

    # 열넓이 조정
    autofit_ColumnSize("O", source_filename)
    autofit_ColumnSize("SO", SO_filename)
    excel_Visible()


def make_Link():
    source_filename = open_ExcelFile("상품수집 링크 정리할 SA 엑셀 선택")
    source_file = pyexcel.get_book(file_name=source_filename)
    link_filename = open_ExcelFile("링크검색할 원장 선택")
    link_sheet = pyexcel.get_sheet(file_name=link_filename, name_columns_by_row=0)
    target_filename = source_filename[:-7] + "SA.link"

    link_list = link_sheet.column[8]
    try:
        SA_sheet = source_file["SA"]
        SA_list = SA_sheet.column[0]
    except:
        print(color_red + "No items!!" + reset)
        return
    with open(target_filename, "w") as f:
        print(f"    ==> Read from : {source_filename}")
        print(f"    ==> link from : {link_filename}")
        print(f"    ==> target file : {target_filename}")
        for i in range(len(SA_list)):
            row_ptr = find_SourceCode(SA_list[i], link_sheet.column["상품코드"])
            if row_ptr != -1:
                print(
                    f"       >> {i} : {SA_list[i]} -> {row_ptr} : {link_sheet[row_ptr, '상품코드']} {link_sheet[row_ptr, '변경상품명']} {link_sheet[row_ptr, '원본경로']}"
                )
                f.write(link_sheet[row_ptr, "원본경로"] + "\n")


def gather_Excel():
    source_excel_filename = open_ExcelFile("정리할 엑셀 선택")
    source_wb = openpyxl.load_workbook(source_excel_filename)
    source_sheet = source_wb.worksheets[0]

    target_excel_filename = source_excel_filename[:-5] + "-t.xlsx"
    print(target_excel_filename)

    target_wb = openpyxl.Workbook(target_excel_filename)
    target_sheet = target_wb.create_sheet()

    row_ptr = 1

    for column in range(1, source_sheet.max_column + 1):
        for row in range(1, source_sheet.max_row + 1):
            if (
                source_sheet.cell(row=row, column=column).value != ""
                and source_sheet.cell(row=row, column=column).value != None
            ):
                print(source_sheet.cell(row=row, column=column).value)
                target_sheet.append([source_sheet.cell(row=row, column=column).value])
                row_ptr += 1
    print(f"\n{row_ptr - 1} item(s) copied")

    target_wb.save(target_excel_filename)


def initialize_Target(mode):
    # 1. 엑셀 로드

    ## 엑셀 오픈
    filename = open_ExcelFile("OSO 처리 대상 엑셀 선택")
    workbook = openpyxl.load_workbook(filename)
    target_option_list_sheet = workbook.worksheets[0]

    ## 대상 로드
    target_option_list = []

    for row in range(target_option_list_sheet.max_row):
        target_option_list.append(
            target_option_list_sheet.cell(row=row + 1, column=1).value
        )
    if mode == "SLG_OSO":
        print("\nHandling Soldout =============================================")
    elif mode == "SLG_SA":
        print("\nHandling to Sell Again =============================================")
    print(f"\n{len(target_option_list)} items will be handled..")
    i = 0
    for target_option in target_option_list:
        i += 1
        print(f"Target Item & Option {i} : {target_option}")
    return filename, target_option_list


def handle_Target(mode, filename, target_option_list):
    workbook = openpyxl.load_workbook(filename)

    ## 시트 정리
    try:
        del workbook["Check"]
    except:
        pass
    workbook.create_sheet("Log", 1)
    Check_sheet = workbook.worksheets[1]
    check_item_list_row = 1

    # 2. 초기화

    # options = Options()
    browser = webdriver.Chrome("./chromedriver.exe", options=Options())

    try:
        ##. 셀러고 이동
        browser.get("http://app.sellergo.net")
        browser.maximize_window()
        # browser.set_window_position(1000, 0, 'current')
        # browser.set_window_size(1560, 1440)
        wait = WebDriverWait(browser, default_Max_wait_seconds)

        ##. 셀러고 로그인
        wait.until(EC.presence_of_element_located((By.NAME, "email"))).send_keys(
            email_id
        )
        wait.until(EC.presence_of_element_located((By.NAME, "password"))).send_keys(pw)
        wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn.btn-success"))
        ).click()
    except:
        print("== Open Error!! ==")
        browser.close()
        return
    ##. 옵션 품절관리 클릭

    browser.implicitly_wait(default_Max_wait_seconds)
    option_soldout_manage_button = browser.find_element_by_xpath(
        "/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[1]/table/tbody/tr[4]/td/table/tbody/tr[1]/td/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div"
    )
    option_soldout_manage_button.click()

    ##. 출력개수 클릭

    option_output_number_button = browser.find_element_by_xpath(
        "/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/table/tbody/tr/td/table/tbody/tr/td[2]/select/option[5]"
    )
    option_output_number_button.click()

    # 3. 루프
    # 셀 복붙, 검색 클릭

    n_target_option_list = len(target_option_list)
    # need_to_check_item = 0
    no_item = 0
    already_checked_item = 0

    for i in range(n_target_option_list):

        # 1단계 : 검색 및 검색 클릭, 전체 체크박스 클릭
        search_box = browser.find_element_by_name("searchBox")

        # browser.implicitly_wait(default_Max_wait_seconds)
        # browser.implicitly_wait(3)
        # time.sleep(0.5)
        time.sleep(1)

        search_text = target_option_list[i]
        search_item_no = target_option_list[i].split()[0]
        search_option = target_option_list[i].split()[
            -(len(target_option_list[i].split()) - 1) :
        ]

        search_box.clear()
        search_box.send_keys(search_text)
        search_button = browser.find_element_by_xpath(
            "/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[3]/button"
        )
        search_button.click()

        # browser.implicitly_wait(default_Max_wait_seconds)
        # browser.implicitly_wait(3)
        # time.sleep(0.5)
        time.sleep(1)

        # if mode == SELL:
        #     all_check_box = browser.find_element_by_css_selector("#gwt-uid-4")
        #     if not all_check_box.is_selected():
        #         all_check_box.click()
        print(
            f"\nProcess {i} / {n_target_option_list}   ============================================="
        )
        print(f"Product No. : {search_item_no} / Search Option : {search_option}")
        search_option_set = set(
            str(" ".join(search_option)).translate(remove_char).split()
        )
        print(f"   ==> Actual Search Text : {search_item_no} {search_option_set}")

        # 품절처리인 경우

        # if mode == SOLDOUT:

        # 2단계 : 검색결과 정리
        soup = BeautifulSoup(browser.page_source, "html.parser")

        searched_items_list = []

        searched_data1 = soup.find_all("tr", class_="flexTable")
        searched_data2 = soup.find_all("tr", class_="flexTable_before_revision")

        for i in range(2):

            if i == 0:
                target_data_list = searched_data1
            else:
                target_data_list = searched_data2
            for j in target_data_list:
                current_td = j.find_all("td")

                id = current_td[0].find("input").attrs["id"]

                item_no = current_td[1].find("div").get_text()
                market_item_no = item_no.replace(search_item_no, "")

                item_name = current_td[2].find("div").get_text()
                item_option = current_td[3].find("div").get_text().strip()
                item_market = current_td[4].find("div").get_text()
                item_status = current_td[6].find("div").get_text()
                searched_items_list.append(
                    dict(
                        id=id,
                        market=item_market,
                        market_item_no=market_item_no,
                        seller_item_no=search_item_no,
                        item_name=item_name,
                        item_option=item_option,
                        item_status=item_status,
                    )
                )
        if len(searched_items_list) - 1 != 0:
            print(f"\n{len(searched_items_list) - 1} items will be checked..\n")
        else:
            print("no item to be checked..\n")
            no_item += 1
        # for searched_item in searched_items_list:
        #     print(searched_item)
        for i in range(1, len(searched_items_list)):
            print(i, " >> ", searched_items_list[i])
        # # 3단계 : 검색결과 처리

        # flag_checked = NOT_CHECKED

        Check_sheet.cell(
            row=check_item_list_row, column=1
        ).value = f"\nProcess {i} / {n_target_option_list}"
        Check_sheet.cell(row=check_item_list_row, column=2).value = search_text
        Check_sheet.cell(
            row=check_item_list_row, column=3
        ).value = f"{len(searched_items_list) - 1} items"
        check_item_list_row += 1

        n_item = 0
        v_checked = 0

        for i in range(1, len(searched_items_list)):

            n_item += 1

            # 판매중인지 확인해서 판매중이 아니면 체크대상 리스트 작성
            # if searched_items_list[i]['item_status'] != '판매중' and flag_checked == NOT_CHECKED:

            # Check_sheet.cell(row=check_item_list_row, column=1).value = search_text

            # if searched_items_list[i]['item_status'] != '판매중':

            Check_sheet.cell(row=check_item_list_row, column=2).value = n_item
            Check_sheet.cell(
                row=check_item_list_row, column=3
            ).value = searched_items_list[i]["market"]
            Check_sheet.cell(
                row=check_item_list_row, column=4
            ).value = searched_items_list[i]["market_item_no"]
            Check_sheet.cell(
                row=check_item_list_row, column=5
            ).value = searched_items_list[i]["seller_item_no"]
            Check_sheet.cell(
                row=check_item_list_row, column=6
            ).value = searched_items_list[i]["item_name"]
            Check_sheet.cell(
                row=check_item_list_row, column=7
            ).value = searched_items_list[i]["item_option"]
            Check_sheet.cell(
                row=check_item_list_row, column=8
            ).value = searched_items_list[i]["item_status"]

            # 옵션이 모두 포함되어 있으면 체크 아니면 체크 안함
            current_option_set = set(
                searched_items_list[i]["item_option"].translate(remove_char).split()
            )
            # current_option_set = set(searched_items_list[i]['item_option'].replace(":", " ").replace("~", " ").replace("(", " ").replace(")", " ").replace("_", " ").replace("-", " ").replace("/", " ").replace("[", " ").replace("]", " ").replace(".", " ").split())
            Check_sheet.cell(row=check_item_list_row, column=9).value = str(
                " ".join(list(current_option_set))
            )
            # search_option_set = set(str(" ".join(search_option)).translate(remove_char).split())
            # search_option_set = set(str(" ".join(search_option)).replace(":", " ").replace("(", " ").replace(")", " ").replace("_", " ").replace("-", " ").replace("/", " ").replace("[", " ").replace("]", " ").replace(".", " ").split())
            Check_sheet.cell(row=check_item_list_row, column=10).value = str(
                " ".join(list(search_option_set))
            )

            if current_option_set & search_option_set == search_option_set:
                if mode == "SLG_OSO":
                    condition1 = "판매중"
                    condition2 = "판매중판매종료"
                elif mode == "SLG_SA":
                    condition1 = "품절"
                    condition2 = "품절판매종료"
                if searched_items_list[i]["item_status"] == condition1:
                    # if searched_items_list[i]['item_status'] == condition1 or searched_items_list[i]['item_status'] == condition2:
                    current_row_checkbox_id = searched_items_list[i]["id"]
                    current_row_checkbox = browser.find_element_by_id(
                        current_row_checkbox_id
                    )
                    current_row_checkbox.click()
                    Check_sheet.cell(row=check_item_list_row, column=1).value = "V"
                    v_checked += 1
            check_item_list_row += 1
        print(f"=> {v_checked} items were checked..")
        check_item_list_row += 1

        # 4단계 : 품절 처리
        if mode == "SLG_OSO":

            soldout_button = browser.find_element_by_xpath(
                "/html/body/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[7]/table/tbody/tr/td/table/tbody/tr/td[1]/button"
            )
            soldout_button.click()
            # browser.implicitly_wait(2)
            # time.sleep(1)
            try:

                # soldout_confirm_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/ html / body / div[2] / div / div / div[2] / button[2]')))
                soldout_confirm_button = browser.find_element_by_xpath(
                    "/ html / body / div[2] / div / div / div[2] / button[2]"
                )
            except:
                # ok_button = browser.find_element_by_xpath("/ html / body / div[2] / div / div / div[2] / button")
                # ok_button = browser.find_element_by_css_selector("body > div.bootbox.modal.fade.bootbox - alert. in > div > div > div.modal - footer > button")
                # ok_button.click()
                # ok_button = browser.find_element_by_css_selector(".btn.btn-primary")
                # ok_button.click()
                wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn.btn-primary"))
                ).click()

                Check_sheet.cell(
                    row=check_item_list_row, column=2
                ).value = "품절처리 실패 및 확인 필요"

                check_item_list_row += 2
                # need_to_check_item += 1
                already_checked_item += 1
            else:
                soldout_confirm_button.click()
                check_item_list_row += 2
            # check_Proceed()
        elif mode == "SLG_SA":

            # 4단계 : 판매 처리
            sell_button = browser.find_element_by_xpath(
                "//*[@id='workingPage']/tbody/tr[2]/td/table/tbody/tr/td[3]/table/tbody/tr[3]/td/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[7]/table/tbody/tr/td/table/tbody/tr/td[3]/button"
            )
            sell_button.click()
            # time.sleep(0.5)
            # time.sleep(1)
            try:
                browser.find_element_by_xpath(
                    "/html/body/div[2]/div/div/div[2]/div/form/input"
                ).send_keys(DEFAULT_STOCK)
            except:
                wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn.btn-primary"))
                ).click()
            else:
                browser.find_element_by_xpath(
                    "/ html / body / div[2] / div / div / div[3] / button[2]"
                ).click()
        if i % 10 == 0:
            workbook.save(filename)
    # Closing
    if mode == "SLG_OSO":
        print(f"\nTotally {n_target_option_list} products were handled.")
        print(f"{no_item} products were blank.")
        print(f"{already_checked_item} products were already checked")
    workbook.save(filename)


def main():

    EXIT = True

    while EXIT:

        # try:
        mode = select_Mode()

        if mode == "CREATE_SB":
            create_SB()
        elif mode == "UPDATE_SB":
            update_SB(mode)
        elif mode == "UPDATE_SO_T":
            update_SO(mode)
        elif mode == "UPDATE_SO_P":
            update_SO(mode)
        elif mode == "GATHER_XLS":
            gather_Excel()
        elif mode == "MAKE_LINK":
            make_Link()
        elif mode == "SLG_OSO" or mode == "SLG_SA":
            filename, target_option_list = initialize_Target(mode)
            handle_Target(mode, filename, target_option_list)
        elif mode == "EXIT":
            excel_Visible()
            EXIT = False
        # except:
        #     pass


if __name__ == "__main__":
    # main

    logger = set_Logger()

    main()
